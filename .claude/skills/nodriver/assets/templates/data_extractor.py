#!/usr/bin/env python3
"""
Comprehensive data extraction template for nodriver.

This template provides production-ready patterns for extracting structured data
from various page layouts including tables, lists, cards, products, images, and
dynamic content. It includes data cleaning, validation, and export functionality.

Key Features:
    - Table data extraction with headers and rows
    - List and card extraction with nested structures
    - Product information scraping
    - Image URL collection with metadata
    - Dynamic content waiting and loading
    - Infinite scroll handling
    - Search and filter operations
    - Data cleaning and validation
    - Multiple export formats (CSV, JSON, Excel)

Example:
    >>> import asyncio
    >>> from data_extractor import DataExtractor
    >>>
    >>> async def main():
    ...     extractor = DataExtractor(tab)
    ...
    ...     # Extract table data
    ...     table_data = await extractor.extract_table(
    ...         "table.data-table",
    ...         header_row=0
    ...     )
    ...
    ...     # Export to CSV
    ...     extractor.export_csv(table_data, "output.csv")
    >>>
    >>> asyncio.run(main())
"""

import asyncio
import csv
import json
import re
import time
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Callable, Union
from urllib.parse import urljoin, urlparse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataValidationError(Exception):
    """Raised when data validation fails."""
    pass


class ExportFormat(Enum):
    """Supported export formats."""
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    JSON_LINES = "jsonl"


@dataclass
class ExtractionResult:
    """Result of a data extraction operation."""
    success: bool
    data: List[Dict[str, Any]] = None
    item_count: int = 0
    extraction_time: float = 0.0
    error: Optional[str] = None
    warnings: List[str] = None

    def __post_init__(self):
        if self.warnings is None:
            self.warnings = []


@dataclass
class ProductInfo:
    """Structured product information."""
    name: str
    price: Optional[float] = None
    description: Optional[str] = None
    url: Optional[str] = None
    image_url: Optional[str] = None
    rating: Optional[float] = None
    in_stock: Optional[bool] = None
    attributes: Dict[str, str] = None

    def __post_init__(self):
        if self.attributes is None:
            self.attributes = {}


class DataValidator:
    """Validates and cleans extracted data."""

    @staticmethod
    def clean_text(text: str) -> str:
        """
        Clean whitespace and normalize text.

        Args:
            text: Raw text to clean

        Returns:
            Cleaned text with normalized whitespace
        """
        if not text:
            return ""

        # Remove extra whitespace
        text = " ".join(text.split())
        # Remove common artifacts
        text = text.strip()
        return text

    @staticmethod
    def extract_number(text: str) -> Optional[float]:
        """
        Extract numeric value from text.

        Args:
            text: Text containing a number

        Returns:
            Extracted float or None if not found
        """
        if not text:
            return None

        # Extract first number (including decimals)
        match = re.search(r'-?\d+\.?\d*', text.replace(',', ''))
        if match:
            try:
                return float(match.group())
            except ValueError:
                return None
        return None

    @staticmethod
    def extract_percentage(text: str) -> Optional[float]:
        """
        Extract percentage value from text.

        Args:
            text: Text containing a percentage

        Returns:
            Percentage as float (0-100) or None
        """
        if not text:
            return None

        match = re.search(r'(\d+\.?\d*)%', text)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                return None
        return None

    @staticmethod
    def normalize_url(url: str, base_url: str = "") -> str:
        """
        Normalize and resolve relative URLs.

        Args:
            url: URL to normalize
            base_url: Base URL for relative URL resolution

        Returns:
            Absolute URL
        """
        if not url:
            return ""

        url = url.strip()

        # Handle relative URLs
        if url.startswith('/'):
            if base_url:
                parsed = urlparse(base_url)
                return f"{parsed.scheme}://{parsed.netloc}{url}"
        elif not url.startswith(('http://', 'https://', '//')):
            if base_url:
                return urljoin(base_url, url)

        return url

    @staticmethod
    def validate_email(email: str) -> bool:
        """
        Validate email address format.

        Args:
            email: Email to validate

        Returns:
            True if valid, False otherwise
        """
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))

    @staticmethod
    def remove_duplicates(
        items: List[Dict[str, Any]],
        key: str
    ) -> List[Dict[str, Any]]:
        """
        Remove duplicate items based on key.

        Args:
            items: List of items
            key: Dictionary key to check for duplicates

        Returns:
            List with duplicates removed
        """
        seen = set()
        unique_items = []

        for item in items:
            value = item.get(key)
            if value not in seen:
                seen.add(value)
                unique_items.append(item)

        return unique_items


class DataExtractor:
    """
    Extract structured data from web pages.

    This class provides methods for extracting various types of data from
    HTML elements including tables, lists, products, and images.
    """

    def __init__(
        self,
        tab: Any,
        base_url: str = "",
        timeout: float = 10.0,
        wait_interval: float = 0.1,
    ):
        """
        Initialize DataExtractor.

        Args:
            tab: nodriver tab object
            base_url: Base URL for relative URL resolution
            timeout: Default timeout for operations
            wait_interval: Interval between polling operations
        """
        self.tab = tab
        self.base_url = base_url
        self.timeout = timeout
        self.wait_interval = wait_interval
        self.validator = DataValidator()

    async def wait_for_load(
        self,
        selector: str,
        timeout: Optional[float] = None
    ) -> bool:
        """
        Wait for element to load and be visible.

        Args:
            selector: CSS selector of element to wait for
            timeout: Override default timeout

        Returns:
            True if element loaded, False if timeout
        """
        timeout = timeout or self.timeout
        start_time = time.time()

        while time.time() - start_time < timeout:
            try:
                elements = await self.tab.select_all(selector)
                if elements:
                    # Check if visible
                    script = """
                    function isVisible(el) {
                        const style = window.getComputedStyle(el);
                        return style.display !== 'none' &&
                               style.visibility !== 'hidden' &&
                               el.offsetParent !== null;
                    }
                    return isVisible(arguments[0]);
                    """
                    is_visible = await self.tab.evaluate(script, elements[0])
                    if is_visible:
                        return True
            except Exception:
                pass

            await asyncio.sleep(self.wait_interval)

        return False

    async def extract_table(
        self,
        selector: str,
        header_row: int = 0,
        clean_data: bool = True,
    ) -> ExtractionResult:
        """
        Extract table data with headers and rows.

        Args:
            selector: CSS selector for table element
            header_row: Index of header row (0-based)
            clean_data: Whether to clean extracted text

        Returns:
            ExtractionResult with table data as list of dicts
        """
        start_time = time.time()

        try:
            # Wait for table to load
            if not await self.wait_for_load(selector):
                return ExtractionResult(
                    success=False,
                    error=f"Table not found: {selector}"
                )

            # Extract table data
            script = """
            function extractTable(tableEl) {
                const headers = [];
                const rows = [];

                // Extract header row
                const headerCells = tableEl.querySelectorAll('thead tr:nth-child(INDEX) td, thead tr:nth-child(INDEX) th');
                if (headerCells.length === 0) {
                    const allRows = tableEl.querySelectorAll('tbody tr, tr');
                    if (allRows.length > INDEX) {
                        const cells = allRows[INDEX].querySelectorAll('td, th');
                        cells.forEach(cell => headers.push(cell.textContent.trim()));
                    }
                } else {
                    headerCells.forEach(cell => headers.push(cell.textContent.trim()));
                }

                // Extract data rows
                const bodyRows = tableEl.querySelectorAll('tbody tr, tr');
                bodyRows.forEach((row, idx) => {
                    if (idx !== INDEX) {
                        const cells = row.querySelectorAll('td, th');
                        const rowData = {};
                        cells.forEach((cell, cellIdx) => {
                            const header = headers[cellIdx] || `Column ${cellIdx}`;
                            rowData[header] = cell.textContent.trim();
                        });
                        if (Object.keys(rowData).length > 0) {
                            rows.push(rowData);
                        }
                    }
                });

                return { headers, rows };
            }

            const tables = document.querySelectorAll('SELECTOR');
            if (tables.length === 0) return { headers: [], rows: [] };

            return extractTable(tables[0]);
            """

            script = script.replace("SELECTOR", selector).replace("INDEX", str(header_row + 1))

            result = await self.tab.evaluate(script)
            rows = result.get('rows', [])

            # Clean data if requested
            if clean_data:
                rows = self._clean_records(rows)

            extraction_time = time.time() - start_time

            return ExtractionResult(
                success=True,
                data=rows,
                item_count=len(rows),
                extraction_time=extraction_time
            )

        except Exception as e:
            extraction_time = time.time() - start_time
            return ExtractionResult(
                success=False,
                extraction_time=extraction_time,
                error=str(e)
            )

    async def extract_list(
        self,
        item_selector: str,
        field_extractors: Dict[str, str],
        clean_data: bool = True,
    ) -> ExtractionResult:
        """
        Extract list or card data using field selectors.

        Args:
            item_selector: CSS selector for each list item
            field_extractors: Dict mapping field names to CSS selectors
                Example: {"name": ".title", "price": ".price"}
            clean_data: Whether to clean extracted text

        Returns:
            ExtractionResult with extracted list items
        """
        start_time = time.time()

        try:
            if not await self.wait_for_load(item_selector):
                return ExtractionResult(
                    success=False,
                    error=f"List items not found: {item_selector}"
                )

            script = """
            function extractList(itemSelector, fieldSelectors) {
                const items = [];
                const itemElements = document.querySelectorAll(itemSelector);

                itemElements.forEach(item => {
                    const itemData = {};
                    for (const [field, selector] of Object.entries(fieldSelectors)) {
                        const el = item.querySelector(selector);
                        itemData[field] = el ? el.textContent.trim() : '';
                    }
                    if (Object.values(itemData).some(v => v)) {
                        items.push(itemData);
                    }
                });

                return items;
            }

            return extractList(ITEM_SELECTOR, FIELD_EXTRACTORS);
            """

            script = script.replace("ITEM_SELECTOR", json.dumps(item_selector))
            script = script.replace("FIELD_EXTRACTORS", json.dumps(field_extractors))

            items = await self.tab.evaluate(script)

            # Clean data if requested
            if clean_data:
                items = self._clean_records(items)

            extraction_time = time.time() - start_time

            return ExtractionResult(
                success=True,
                data=items,
                item_count=len(items),
                extraction_time=extraction_time
            )

        except Exception as e:
            extraction_time = time.time() - start_time
            return ExtractionResult(
                success=False,
                extraction_time=extraction_time,
                error=str(e)
            )

    async def extract_products(
        self,
        item_selector: str,
        name_selector: str,
        price_selector: Optional[str] = None,
        image_selector: Optional[str] = None,
        rating_selector: Optional[str] = None,
        url_selector: Optional[str] = None,
    ) -> ExtractionResult:
        """
        Extract product information from list.

        Args:
            item_selector: CSS selector for each product item
            name_selector: CSS selector for product name
            price_selector: CSS selector for price
            image_selector: CSS selector for image
            rating_selector: CSS selector for rating
            url_selector: CSS selector for product URL

        Returns:
            ExtractionResult with product data
        """
        start_time = time.time()

        try:
            if not await self.wait_for_load(item_selector):
                return ExtractionResult(
                    success=False,
                    error=f"Products not found: {item_selector}"
                )

            script = """
            function extractProducts(selectors) {
                const products = [];
                const items = document.querySelectorAll(selectors.item);

                items.forEach(item => {
                    const product = {
                        name: '',
                        price: null,
                        image_url: '',
                        rating: null,
                        url: ''
                    };

                    // Extract name
                    if (selectors.name) {
                        const nameEl = item.querySelector(selectors.name);
                        product.name = nameEl ? nameEl.textContent.trim() : '';
                    }

                    // Extract price
                    if (selectors.price) {
                        const priceEl = item.querySelector(selectors.price);
                        if (priceEl) {
                            const priceText = priceEl.textContent;
                            const match = priceText.match(/\d+\.?\d*/);
                            product.price = match ? parseFloat(match[0]) : null;
                        }
                    }

                    // Extract image
                    if (selectors.image) {
                        const imgEl = item.querySelector(selectors.image);
                        if (imgEl) {
                            product.image_url = imgEl.src || imgEl.getAttribute('data-src') || '';
                        }
                    }

                    // Extract rating
                    if (selectors.rating) {
                        const ratingEl = item.querySelector(selectors.rating);
                        if (ratingEl) {
                            const ratingText = ratingEl.textContent;
                            const match = ratingText.match(/\d+\.?\d*/);
                            product.rating = match ? parseFloat(match[0]) : null;
                        }
                    }

                    // Extract URL
                    if (selectors.url) {
                        const linkEl = item.querySelector(selectors.url);
                        product.url = linkEl ? linkEl.href : '';
                    }

                    if (product.name) {
                        products.push(product);
                    }
                });

                return products;
            }

            return extractProducts(SELECTORS);
            """

            selectors = {
                'item': item_selector,
                'name': name_selector,
                'price': price_selector,
                'image': image_selector,
                'rating': rating_selector,
                'url': url_selector,
            }

            script = script.replace("SELECTORS", json.dumps(selectors))
            products = await self.tab.evaluate(script)

            # Convert to ProductInfo objects and normalize URLs
            product_list = []
            for product in products:
                product['image_url'] = self.validator.normalize_url(
                    product.get('image_url', ''),
                    self.base_url
                )
                product['url'] = self.validator.normalize_url(
                    product.get('url', ''),
                    self.base_url
                )
                product_list.append(product)

            extraction_time = time.time() - start_time

            return ExtractionResult(
                success=True,
                data=product_list,
                item_count=len(product_list),
                extraction_time=extraction_time
            )

        except Exception as e:
            extraction_time = time.time() - start_time
            return ExtractionResult(
                success=False,
                extraction_time=extraction_time,
                error=str(e)
            )

    async def extract_images(
        self,
        selector: str,
        attr: str = "src",
        extract_alt: bool = True,
    ) -> ExtractionResult:
        """
        Extract image URLs with optional metadata.

        Args:
            selector: CSS selector for images
            attr: Image attribute to extract URL from (src, data-src, etc.)
            extract_alt: Whether to include alt text

        Returns:
            ExtractionResult with image data
        """
        start_time = time.time()

        try:
            script = """
            function extractImages(selector, attr, extractAlt) {
                const images = [];
                const elements = document.querySelectorAll(selector);

                elements.forEach(img => {
                    let url = img.getAttribute(attr);
                    if (!url) url = img.getAttribute('src');
                    if (!url) url = img.getAttribute('data-src');

                    if (url) {
                        const item = { url: url };
                        if (extractAlt) {
                            item.alt = img.getAttribute('alt') || '';
                            item.title = img.getAttribute('title') || '';
                        }
                        images.push(item);
                    }
                });

                return images;
            }

            return extractImages(SELECTOR, ATTR, EXTRACT_ALT);
            """

            script = script.replace("SELECTOR", json.dumps(selector))
            script = script.replace("ATTR", json.dumps(attr))
            script = script.replace("EXTRACT_ALT", str(extract_alt).lower())

            images = await self.tab.evaluate(script)

            # Normalize URLs
            for image in images:
                image['url'] = self.validator.normalize_url(
                    image['url'],
                    self.base_url
                )

            extraction_time = time.time() - start_time

            return ExtractionResult(
                success=True,
                data=images,
                item_count=len(images),
                extraction_time=extraction_time
            )

        except Exception as e:
            extraction_time = time.time() - start_time
            return ExtractionResult(
                success=False,
                extraction_time=extraction_time,
                error=str(e)
            )

    async def handle_infinite_scroll(
        self,
        item_selector: str,
        scroll_pause: float = 2.0,
        max_scrolls: int = 10,
        pixels_per_scroll: int = 500,
    ) -> ExtractionResult:
        """
        Handle infinite scroll pagination.

        Args:
            item_selector: CSS selector for items to count
            scroll_pause: Time to wait between scrolls (seconds)
            max_scrolls: Maximum number of scroll operations
            pixels_per_scroll: Pixels to scroll per operation

        Returns:
            ExtractionResult with item count progress
        """
        start_time = time.time()
        previous_count = 0
        no_change_count = 0

        try:
            for scroll_num in range(max_scrolls):
                # Count current items
                script = f"return document.querySelectorAll('{item_selector}').length;"
                current_count = await self.tab.evaluate(script)

                if current_count == previous_count:
                    no_change_count += 1
                    if no_change_count >= 2:
                        break  # No new items loaded
                else:
                    no_change_count = 0

                # Scroll down
                scroll_script = f"""
                window.scrollBy(0, {pixels_per_scroll});
                """
                await self.tab.evaluate(scroll_script)

                # Wait for new content to load
                await asyncio.sleep(scroll_pause)

                previous_count = current_count

            extraction_time = time.time() - start_time

            return ExtractionResult(
                success=True,
                item_count=current_count,
                extraction_time=extraction_time,
                data=[{"scroll_count": scroll_num, "items_loaded": current_count}]
            )

        except Exception as e:
            extraction_time = time.time() - start_time
            return ExtractionResult(
                success=False,
                extraction_time=extraction_time,
                error=str(e)
            )

    async def search_and_filter(
        self,
        search_selector: str,
        search_term: str,
        filter_selectors: Optional[Dict[str, str]] = None,
        wait_for: str = ".results",
    ) -> ExtractionResult:
        """
        Perform search and optional filtering.

        Args:
            search_selector: CSS selector for search input
            search_term: Term to search for
            filter_selectors: Dict of filter element selectors
                Example: {"category": ".filter-category"}
            wait_for: Selector to wait for after search

        Returns:
            ExtractionResult with operation status
        """
        start_time = time.time()

        try:
            # Find and fill search input
            search_elements = await self.tab.select_all(search_selector)
            if not search_elements:
                return ExtractionResult(
                    success=False,
                    error=f"Search input not found: {search_selector}"
                )

            search_input = search_elements[0]
            await search_input.clear()
            await search_input.type(search_term)

            # Apply filters if provided
            if filter_selectors:
                for filter_name, filter_selector in filter_selectors.items():
                    filters = await self.tab.select_all(filter_selector)
                    if filters:
                        await filters[0].click()
                        await asyncio.sleep(0.5)

            # Wait for results
            if wait_for and not await self.wait_for_load(wait_for):
                return ExtractionResult(
                    success=False,
                    error=f"Results not loaded: {wait_for}"
                )

            extraction_time = time.time() - start_time

            return ExtractionResult(
                success=True,
                extraction_time=extraction_time,
                data=[{"search_term": search_term, "filters": filter_selectors}]
            )

        except Exception as e:
            extraction_time = time.time() - start_time
            return ExtractionResult(
                success=False,
                extraction_time=extraction_time,
                error=str(e)
            )

    def export_csv(
        self,
        data: List[Dict[str, Any]],
        filepath: Union[str, Path],
    ) -> bool:
        """
        Export data to CSV file.

        Args:
            data: List of dictionaries to export
            filepath: Output file path

        Returns:
            True if successful, False otherwise
        """
        if not data:
            return False

        try:
            filepath = Path(filepath)
            filepath.parent.mkdir(parents=True, exist_ok=True)

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)

            return True
        except Exception:
            return False

    def export_json(
        self,
        data: List[Dict[str, Any]],
        filepath: Union[str, Path],
        pretty: bool = True,
    ) -> bool:
        """
        Export data to JSON file.

        Args:
            data: List of dictionaries to export
            filepath: Output file path
            pretty: Whether to pretty-print JSON

        Returns:
            True if successful, False otherwise
        """
        try:
            filepath = Path(filepath)
            filepath.parent.mkdir(parents=True, exist_ok=True)

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2 if pretty else None, ensure_ascii=False)

            return True
        except Exception:
            return False

    def export_jsonl(
        self,
        data: List[Dict[str, Any]],
        filepath: Union[str, Path],
    ) -> bool:
        """
        Export data to JSONL file (one JSON object per line).

        Args:
            data: List of dictionaries to export
            filepath: Output file path

        Returns:
            True if successful, False otherwise
        """
        try:
            filepath = Path(filepath)
            filepath.parent.mkdir(parents=True, exist_ok=True)

            with open(filepath, 'w', encoding='utf-8') as f:
                for item in data:
                    f.write(json.dumps(item, ensure_ascii=False) + '\n')

            return True
        except Exception:
            return False

    def export_excel(
        self,
        data: List[Dict[str, Any]],
        filepath: Union[str, Path],
        sheet_name: str = "Data",
    ) -> bool:
        """
        Export data to Excel file.

        Args:
            data: List of dictionaries to export
            filepath: Output file path
            sheet_name: Name of worksheet

        Returns:
            True if successful, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl not installed. Install with: pip install openpyxl"
            )

        if not data:
            return False

        try:
            filepath = Path(filepath)
            filepath.parent.mkdir(parents=True, exist_ok=True)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

            wb.save(filepath)
            return True

        except Exception:
            return False

    def export(
        self,
        data: List[Dict[str, Any]],
        filepath: Union[str, Path],
        format: ExportFormat = ExportFormat.JSON,
    ) -> bool:
        """
        Export data in specified format.

        Args:
            data: Data to export
            filepath: Output file path
            format: Export format

        Returns:
            True if successful, False otherwise
        """
        if format == ExportFormat.CSV:
            return self.export_csv(data, filepath)
        elif format == ExportFormat.JSON:
            return self.export_json(data, filepath)
        elif format == ExportFormat.JSONL:
            return self.export_jsonl(data, filepath)
        elif format == ExportFormat.EXCEL:
            return self.export_excel(data, filepath)
        else:
            raise ValueError(f"Unsupported format: {format}")

    @staticmethod
    def _clean_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Clean all text fields in records.

        Args:
            records: Records to clean

        Returns:
            Records with cleaned text fields
        """
        validator = DataValidator()
        cleaned = []

        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, str):
                    cleaned_record[key] = validator.clean_text(value)
                else:
                    cleaned_record[key] = value
            cleaned.append(cleaned_record)

        return cleaned


async def main():
    """Example usage of DataExtractor."""
    print("Data Extraction Template for nodriver")
    print("=" * 60)
    print("\nUsage examples:\n")

    print("1. Extract table data:")
    print("   result = await extractor.extract_table('table.data-table')")
    print("   if result.success:")
    print("       extractor.export_csv(result.data, 'output.csv')")

    print("\n2. Extract list items:")
    print("   result = await extractor.extract_list(")
    print("       '.product-item',")
    print("       {'name': '.title', 'price': '.price'}")
    print("   )")

    print("\n3. Extract product data:")
    print("   result = await extractor.extract_products(")
    print("       '.product',")
    print("       '.product-name',")
    print("       price_selector='.product-price',")
    print("       image_selector='img.product-image'")
    print("   )")

    print("\n4. Extract images:")
    print("   result = await extractor.extract_images('img.gallery')")
    print("   extractor.export_json(result.data, 'images.json')")

    print("\n5. Handle infinite scroll:")
    print("   result = await extractor.handle_infinite_scroll(")
    print("       '.product-item',")
    print("       scroll_pause=2.0,")
    print("       max_scrolls=10")
    print("   )")

    print("\n6. Search and filter:")
    print("   result = await extractor.search_and_filter(")
    print("       'input#search',")
    print("       'laptop',")
    print("       filter_selectors={'category': '.filter-cat'}")
    print("   )")

    print("\n7. Export in multiple formats:")
    print("   extractor.export_csv(data, 'data.csv')")
    print("   extractor.export_json(data, 'data.json')")
    print("   extractor.export_excel(data, 'data.xlsx')")
    print("   extractor.export_jsonl(data, 'data.jsonl')")

    print("\n8. Data validation and cleaning:")
    print("   price = DataValidator.extract_number('$19.99')")
    print("   email_valid = DataValidator.validate_email('test@example.com')")
    print("   url = DataValidator.normalize_url('/products', base_url)")
    print("   clean_text = DataValidator.clean_text('  Extra  spaces  ')")


if __name__ == "__main__":
    asyncio.run(main())
